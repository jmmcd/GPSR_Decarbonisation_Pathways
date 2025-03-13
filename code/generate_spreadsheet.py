from openpyxl import Workbook
import re
import pandas as pd
import numpy as np

# a script to convert an equation to an Excel formula
# Claude.ai generated most of this

def convert_equation_to_excel(equation, var_columns):
    """Convert algebraic equation to Excel formula using cell references"""
    excel_eq = equation
    for var, col in var_columns.items():
        excel_eq = re.sub(r'\b' + var + r'\b', f'${col}$2', excel_eq)
    return f'={excel_eq}'

def create_excel_with_equations(x_vars, y_equations_sets, typical_values, output_file='equations.xlsx'):
    wb = Workbook()
    ws = wb.active
    
    # Create column mapping for variables
    var_columns = {var: chr(65 + i) for i, var in enumerate(x_vars)}
    
    # Write headers
    col = 1
    for var in x_vars:
        ws.cell(row=1, column=col, value=var)
        ws.cell(row=2, column=col, value=typical_values[var])
        col += 1
    
    # Write Y equations
    for y_var in y_equations_sets:
        ws.cell(row=1, column=col, value=y_var)
        
        row = 2
        for eqn in y_equations_sets[y_var]:
            # Convert equation and write formula
            print(eqn)
            excel_formula = convert_equation_to_excel(eqn, var_columns)
            ws.cell(row=row, column=col, value=excel_formula)
            row += 1
        col += 1
    
    wb.save(output_file)

### X
df = pd.read_csv('../outputs/data_XY.csv', index_col=0)

renames = {
    'Scenario': 'Scenario', 
    'SOW': 'SOW', 
    'Temperature Limit': 'Temp_Limit', 
    'Delayed Action': 'Delay', 
    'GDP': 'GDP',
    'Population': 'Pop', 
    'Other Energy Service Demand Drivers': 'Other_ESD_Drivers',
    'Social discount rate': 'SDR',
    'Elasticity of energy service demand to its own driver': 'Elast_ESD_Driver',
    'Elasticity of energy service demand to its own price': 'Elast_ESD_Price',
    'CO2 Storage Potential': 'CO2_Storage_Poten', 
    'Wind Potential': 'Wind_Poten', 
    'Solar Potential': 'Solar_Poten',
    'Biomass Potential': 'Biomass_Poten', 
    'Oil & Gas Potential': 'Oil_Gas_Poten', 
    'Solar PV  Investment Cost': 'Solar_PV_Inv_Cost',
    'Wind Investment Cost': 'Wind_Inv_Cost', 
    'Bioenergy with CCS Specific Investment Cost': 'Bioenergy_CCS_Inv_Cost',
    'Other technologies costs': 'Other_Tech_Cost', 
    'Forcing of non-energy emissions': 'Forcing',
    'Land Use, Land Use Change and Forestry Sinks': 'Land_Sinks',
    'Climate Sensitivity (in oC)': 'Clim_Sens', 
    'Year': 'Year',
    'Temperature change (oC)': 'Temp_Change',
    'Radiative Forcing (W/sqm)': 'Rad_Forcing',
    'CO2 concentrations (PPM)': 'CO2_Conc',
    'CH4 Concentrations (ppb)': 'CH4_Conc', 
    'N2O concentrations (ppb)': 'N20_Conc', 
    'Carbon (Gt)': 'Carbon',
    'CO2 (Gt-eq)': 'CO2', 
    'Marginal CO2 cost (USD/t)': 'Marg_CO2_Cost',
    'Global Total Electricity Supply (EJ/yr.)': 'GSupply',
    'Global Electricty Supply from Bioenergy (EJ/yr.)': 'GSupply_Bioenergy',
    'Global Electricity Supply from Fossil Fuels (EJ/yr.)': 'GSupply_Fossil',
    'Global Electricity Supply from Geothermal (EJ/yr.)': 'GSupply_Geothermal',
    'Global Electricity Supply from Nuclear (EJ/yr.)': 'GSupply_Nuclear',
    'Global Electricity Supply from Solar (EJ/yr.)': 'GSupply_Solar',
    'Global Electricity Supply from Tidal/Waves (EJ/yr.)': 'GSupply_Tidal_Waves',
    'Global Electricity Supply from Wind (EJ/yr.)': 'GSupply_Wind',
    'Global Total Primary Energy Consumption (EJ/yr.)': 'GConsumption',
    'Global Primary Energy Consumption of Fossil Fuels (EJ/yr.)': 'GConsumption_Fossil',
    'Global Primary Energy Consumption of Nuclear energy (EJ/yr.)': 'GConsumption_Nuclear',
    'Global Primary Energy Consumption of Renewable energy (EJ/yr.)': 'GConsumption_Renewable',
    'Annual Total Global Energy System Cost (Trillion USD)': 'GCost'
}

df.rename(renames, axis=1, inplace=True)
x_vars = ['GDP', 'Pop',
       'Other_ESD_Drivers', 'SDR', 'Elast_ESD_Driver', 'Elast_ESD_Price',
       'CO2_Storage_Poten', 'Wind_Poten', 'Solar_Poten', 'Biomass_Poten',
       'Oil_Gas_Poten', 'Solar_PV_Inv_Cost', 'Wind_Inv_Cost',
       'Bioenergy_CCS_Inv_Cost', 'Other_Tech_Cost', 'Forcing', 'Land_Sinks',
       'Clim_Sens']



scenarios = ['BASE_SSP2', '1p5c_OS_SSP2', '2C_SSP2', '2C_SSP2_DA30']

for scenario in scenarios:

    # TODO: make the typical values and the equations depend on the scenario, ie read different
    # equation files and subset the X differently
    df_tmp = df[df['Scenario'] == scenario][x_vars]
    dfm = df_tmp.mean(axis=0)
    # print(dfm)
    typical_values = {v: dfm[v] for v in x_vars}
    print(typical_values)

    ### Y

    y_equations = {}
    y_vars = ['GCost', 'GConsumption', 'CO2']

    for y_var in y_vars:
        df_y = pd.read_csv(f'../outputs/equations_sr_{scenario}_{y_var}.csv')
        eqns = list(df_y['equation'])
        y_equations[y_var] = eqns

    print(y_equations)
    create_excel_with_equations(x_vars, y_equations, typical_values, output_file=f'../outputs/equations_{scenario}.xlsx')